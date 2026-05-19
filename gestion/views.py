from datetime import date,timedelta
from django.contrib import messages  # type: ignore
from django.contrib.auth.decorators import login_required # type: ignore
from django.core.paginator import Paginator # type: ignore
from django.db.models import Count, Sum,Q # type: ignore
from django.http import HttpResponseForbidden,HttpResponse # type: ignore
import pandas as pd # type: ignore
from reportlab.pdfgen import canvas # type: ignore
from django.shortcuts import get_object_or_404, redirect, render # type: ignore
from django.utils import timezone # type: ignore
from django.core.mail import send_mail # type: ignore
from .decorators import role_required
from django.shortcuts import render # type: ignore
from django.contrib.auth.models import User # type: ignore
from django.contrib.auth.forms import PasswordResetForm # type: ignore
import deepl # type: ignore
from django.conf import settings # type: ignore
from openpyxl import Workbook # type: ignore
from reportlab.lib.pagesizes import A4, landscape # type: ignore
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer,Table,TableStyle,PageBreak  # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.lib.styles import getSampleStyleSheet # type: ignore
from collections import defaultdict
from decimal import Decimal, ROUND_UP
from django.utils.safestring import mark_safe # type: ignore
from urllib.parse import quote
from django.db import transaction
from django.utils.dateparse import parse_date




from .forms import (
    ActiviteCulturelleForm,
    AffiliationAssociationForm,
    AssociationForm,
    BeneficiaireFederationForm,
    CotisationAssociationMensuelleForm,
    DocumentForm,
    DossierRapatriementForm,
    PersonneForm,
    TransactionForm,
    InformationForm,
    ContactPublicForm,
    ContributionDecesAssociationForm,
    AdhesionForm,
    InvitationUtilisateurForm,
    BureauMembreForm,
    BureauForm, 
    PosteForm,
    AssociationUpdateForm,
    PaiementAssociationForm,
)
from .models import (
    ActiviteCulturelle,
    AffiliationAssociation,
    Association,
    BeneficiaireFederation,
    CotisationAssociationMensuelle,
    Document,
    DossierRapatriement,
    Personne,
    Transaction,
    Information,
    Notification,
    Entite,
    ActiviteCulturelle,
    ContributionDecesAssociation,
    Adhesion,
    MessageContact,
    Bureau,
    BureauMembre,
    Poste,
    PaiementAssociation,
)
from .utils import (
                get_user_entites, user_has_role, user_has_permission,notifier_users,
                generer_numero_personne,get_resume_financier_associations,)

from .emails import (
    envoyer_rappel_cotisation as envoyer_email_rappel_cotisation,
    envoyer_confirmation_paiement,
    envoyer_notification_activite,
    envoyer_email_historise,
)
from .models import HistoriqueEmail


MONTH_LABELS = {
    1: 'Jan', 2: 'Fév', 3: 'Mar', 4: 'Avr', 5: 'Mai', 6: 'Juin',
    7: 'Juil', 8: 'Août', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Déc'
}


def paginate(request, queryset, per_page=15):
    paginator = Paginator(queryset, per_page)
    return paginator.get_page(request.GET.get('page'))



@login_required
def dashboard(request):
    entites = get_user_entites(request.user)
    federations = entites.filter(type="federation")
    associations_entites = entites.filter(type="association")

    federation = federations.first()
    association_entite = associations_entites.first()

    if federation and user_has_permission(request.user, federation, "dashboard_view"):
        current_year = date.today().year
        current_month = date.today().month

        associations_count = Association.objects.filter(federation=federation).count()

        beneficiaires_count = BeneficiaireFederation.objects.filter(
            affiliation__federation=federation,
            actif=True,
            personne__statut="actif"
        ).count()

        cotisations_mois = CotisationAssociationMensuelle.objects.filter(
            affiliation__federation=federation,
            annee=current_year,
            mois=current_month
        )

        montant_attendu = cotisations_mois.aggregate(total=Sum("montant_attendu"))["total"] or 0
        montant_paye = cotisations_mois.aggregate(total=Sum("montant_paye"))["total"] or 0
        reste_a_payer = montant_attendu - montant_paye

        cotisations_payees = cotisations_mois.filter(statut="paye").count()
        cotisations_impayees = cotisations_mois.exclude(statut="paye").count()

        taux_paiement = round((montant_paye / montant_attendu) * 100, 2) if montant_attendu else 0

        rapatriements_ouverts = DossierRapatriement.objects.filter(
            affiliation__federation=federation,
            statut__in=["ouvert", "en_cours"]
        ).count()

        activites_a_venir = ActiviteCulturelle.objects.filter(
            entite=federation,
            date__gte=date.today()
        ).order_by("date")[:5]

        informations_recentes = Information.objects.filter(
            entite=federation,
            statut="publie"
        ).order_by("-created_at")[:5]

        cotisations_par_mois = CotisationAssociationMensuelle.objects.filter(
            affiliation__federation=federation,
            annee=current_year
        ).values("mois").annotate(total=Sum("montant_paye")).order_by("mois")

        chart_labels = [MONTH_LABELS[row["mois"]] for row in cotisations_par_mois]
        chart_values = [float(row["total"] or 0) for row in cotisations_par_mois]

        mauvais_payeurs = CotisationAssociationMensuelle.objects.filter(
            affiliation__federation=federation,
            statut__in=["impaye", "partiel", "en_attente"]
        ).values(
            "affiliation__association__entite__nom"
        ).annotate(
            total=Count("id")
        ).order_by("-total")[:5]

        retard_labels = [row["affiliation__association__entite__nom"] for row in mauvais_payeurs]
        retard_values = [row["total"] for row in mauvais_payeurs]

        alertes_cotisations = CotisationAssociationMensuelle.objects.filter(
            affiliation__federation=federation,
            annee=current_year,
            mois=current_month,
            statut__in=["impaye", "en_attente", "partiel"]
        ).select_related("affiliation__association__entite")[:5]

        alertes_rapatriements = DossierRapatriement.objects.filter(
            affiliation__federation=federation,
            statut__in=["ouvert", "en_cours"]
        ).select_related("personne", "affiliation__association__entite")[:5]

        limite = date.today() + timedelta(days=7)

        alertes_activites = ActiviteCulturelle.objects.filter(
            entite=federation,
            date__gte=date.today(),
            date__lte=limite
        ).order_by("date")[:5]

        return render(request, "dashboard_federation.html", {
            "associations_count": associations_count,
            "beneficiaires_count": beneficiaires_count,
            "cotisations_payees": cotisations_payees,
            "cotisations_impayees": cotisations_impayees,
            "montant_attendu": montant_attendu,
            "montant_paye": montant_paye,
            "reste_a_payer": reste_a_payer,
            "taux_paiement": taux_paiement,
            "rapatriements_ouverts": rapatriements_ouverts,
            "activites_a_venir": activites_a_venir,
            "informations_recentes": informations_recentes,
            "chart_labels": chart_labels,
            "chart_values": chart_values,
            "retard_labels": retard_labels,
            "retard_values": retard_values,
            "alertes_cotisations": alertes_cotisations,
            "alertes_rapatriements": alertes_rapatriements,
            "alertes_activites": alertes_activites,
            "current_year": current_year,
            "current_month": current_month,
        })

    if association_entite:
        association = get_object_or_404(
            Association.objects.select_related("entite", "federation"),
            entite=association_entite
        )

        affiliation = association.affiliations.filter(statut="active").first()

        membres_count = Personne.objects.filter(
            adhesions__association=association
        ).distinct().count()

        nouveaux_membres = Personne.objects.filter(
            adhesions__association=association,
            adhesions__created_at__gte=date.today().replace(day=1)
        ).distinct().count()

        beneficiaires_count = BeneficiaireFederation.objects.filter(
            affiliation__association=association,
            actif=True,
            personne__statut="actif"
        ).count()

        cotisation_mois = None
        cotisation_status = "Aucune cotisation"
        montant_attendu = 0
        montant_paye = 0
        reste_a_payer = 0

        if affiliation:
            cotisation_mois = CotisationAssociationMensuelle.objects.filter(
                affiliation=affiliation,
                annee=date.today().year,
                mois=date.today().month
            ).first()

            if cotisation_mois:
                montant_attendu = cotisation_mois.montant_attendu
                montant_paye = cotisation_mois.montant_paye
                reste_a_payer = montant_attendu - montant_paye
                cotisation_status = cotisation_mois.get_statut_display()

        historique = []
        chart_labels = []
        chart_values = []

        if affiliation:
            historique = CotisationAssociationMensuelle.objects.filter(
                affiliation=affiliation
            ).order_by("-annee", "-mois")[:12]

            historique_reverse = list(reversed(historique))
            chart_labels = [f"{c.mois}/{c.annee}" for c in historique_reverse]
            chart_values = [float(c.montant_paye or 0) for c in historique_reverse]

        activites_a_venir = ActiviteCulturelle.objects.filter(
            entite=association.federation,
            date__gte=date.today()
        ).order_by("date")[:5]

        informations_recentes = Information.objects.filter(
            entite=association.federation,
            statut="publie"
        ).order_by("-created_at")[:5]

        documents_recents = Document.objects.filter(
            Q(entite=association.entite) | Q(entite=association.federation)
        ).order_by("-created_at")[:5]

        return render(request, "dashboard_association.html", {
            "association": association,
            "affiliation": affiliation,
            "membres_count": membres_count,
            "nouveaux_membres": nouveaux_membres,
            "beneficiaires_count": beneficiaires_count,
            "cotisation_mois": cotisation_mois,
            "cotisation_status": cotisation_status,
            "montant_attendu": montant_attendu,
            "montant_paye": montant_paye,
            "reste_a_payer": reste_a_payer,
            "historique": historique,
            "activites_a_venir": activites_a_venir,
            "informations_recentes": informations_recentes,
            "documents_recents": documents_recents,
            "chart_labels": chart_labels,
            "chart_values": chart_values,
        })

    return HttpResponseForbidden("Aucune entité accessible.")

@login_required
def alertes_liste(request):
    federations = get_user_entites(request.user).filter(type='federation')
    today = date.today()
    current_year = today.year
    current_month = today.month

    cotisations = CotisationAssociationMensuelle.objects.filter(
        affiliation__federation__in=federations,
        annee=current_year,
        mois=current_month,
        statut__in=['impaye', 'en_attente', 'partiel']
    ).select_related('affiliation__association__entite')

    rapatriements = DossierRapatriement.objects.filter(
        affiliation__federation__in=federations,
        statut__in=['ouvert', 'en_cours']
    ).select_related('personne', 'affiliation__association__entite')

    activites = ActiviteCulturelle.objects.filter(
        entite__in=federations,
        date__gte=today,
        date__lte=today + timedelta(days=7)
    ).order_by('date')

    return render(request, 'alertes/liste.html', {
        'cotisations': cotisations,
        'rapatriements': rapatriements,
        'activites': activites,
    })


@login_required
def associations_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "associations_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = Association.objects.filter(federation__in=federations).select_related("entite", "federation")

    ville = request.GET.get("ville")
    if ville:
        qs = qs.filter(ville__icontains=ville)

    q = request.GET.get("q")
    if q:
        qs = qs.filter(entite__nom__icontains=q)

    page_obj = paginate(request, qs.order_by("entite__nom"))
    return render(request, "associations/liste.html", {
        "page_obj": page_obj,
        "q": q,
    })


@login_required
def association_detail(request, pk):
    association = get_object_or_404(
        Association.objects.select_related("entite", "federation"),
        pk=pk,
    )

    can_view = (
        user_has_permission(request.user, association.federation, "associations_view")
        or user_has_permission(request.user, association.entite, "association_view_own")
    )

    if not can_view:
        return HttpResponseForbidden("Accès refusé.")

    affiliation = association.affiliations.select_related("federation").first()
    cotisations = affiliation.cotisations_mensuelles.all()[:12] if affiliation else []

    beneficiaires = BeneficiaireFederation.objects.filter(
        affiliation__association=association,
        actif=True,
    ).select_related("personne")[:10]

    return render(request, "associations/detail.html", {
        "association": association,
        "cotisations": cotisations,
        "beneficiaires": beneficiaires,
    })


@login_required
def association_create(request):
    federations = get_user_entites(request.user).filter(type='federation')

    if not federations.exists():
        messages.error(request, "Aucune fédération accessible pour votre compte.")
        return redirect('associations_liste')

    federation = federations.first()
    if not user_has_role(request.user, federation, ['admin', 'manager']):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == 'POST':
        form = AssociationForm(request.POST)
        form.fields['federation'].queryset = federations

        if form.is_valid():
            form.save()
            messages.success(request, 'Association créée avec succès.')
            return redirect('associations_liste')
    else:
        form = AssociationForm()
        form.fields['federation'].queryset = federations

    return render(request, 'associations/form.html', {
        'form': form,
        'title': 'Ajouter une association'
    })

@login_required
def association_update(request, pk):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "associations_edit"):
        return HttpResponseForbidden("Accès refusé.")

    association = get_object_or_404(
        Association.objects.select_related("entite", "federation"),
        pk=pk,
        federation__in=federations
    )

    if request.method == "POST":
        form = AssociationUpdateForm(
            request.POST,
            request.FILES,
            instance=association
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Association modifiée avec succès.")
            return redirect("association_detail", pk=association.pk)
    else:
        form = AssociationUpdateForm(instance=association)

    return render(request, "associations/form.html", {
        "form": form,
        "title": "Modifier une association"
    })

@login_required
def association_delete(request, pk):
    federations = get_user_entites(request.user).filter(type='federation')
    association = get_object_or_404(Association, pk=pk, federation__in=federations)
    if request.method == 'POST':
        association.delete()
        messages.success(request, 'Association supprimée.')
        return redirect('associations_liste')
    return render(request, 'associations/supprimer.html', {'object': association, 'title': 'Supprimer une association'})


@login_required
def affiliations_liste(request):
    federations = get_user_entites(request.user).filter(type='federation')
    qs = AffiliationAssociation.objects.filter(federation__in=federations).select_related('association__entite', 'federation')
    page_obj = paginate(request, qs.order_by('-date_affiliation'))
    return render(request, 'affiliations/liste.html', {'page_obj': page_obj})


@login_required
def affiliation_create(request):
    federations = get_user_entites(request.user).filter(type='federation')
    if request.method == 'POST':
        form = AffiliationAssociationForm(request.POST)
        form.fields['federation'].queryset = federations
        form.fields['association'].queryset = Association.objects.filter(federation__in=federations)
        if form.is_valid():
            form.save()
            messages.success(request, 'Affiliation créée avec succès.')
            return redirect('affiliations_liste')
    else:
        form = AffiliationAssociationForm()
        form.fields['federation'].queryset = federations
        form.fields['association'].queryset = Association.objects.filter(federation__in=federations)
    return render(request, 'affiliations/form.html', {'form': form, 'title': 'Ajouter une affiliation'})


@login_required
def affiliation_update(request, pk):
    federations = get_user_entites(request.user).filter(type='federation')
    affiliation = get_object_or_404(AffiliationAssociation, pk=pk, federation__in=federations)
    if request.method == 'POST':
        form = AffiliationAssociationForm(request.POST, instance=affiliation)
        form.fields['federation'].queryset = federations
        form.fields['association'].queryset = Association.objects.filter(federation__in=federations)
        if form.is_valid():
            form.save()
            messages.success(request, 'Affiliation modifiée avec succès.')
            return redirect('affiliations_liste')
    else:
        form = AffiliationAssociationForm(instance=affiliation)
        form.fields['federation'].queryset = federations
        form.fields['association'].queryset = Association.objects.filter(federation__in=federations)
    return render(request, 'affiliations/form.html', {'form': form, 'title': 'Modifier une affiliation'})


@login_required
def personnes_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations = get_user_entites(request.user).filter(type="association")

    qs = Personne.objects.none()

    federation = federations.first()

    if federation and user_has_permission(request.user, federation, "membres_view_all"):
        qs = Personne.objects.filter(
            adhesions__association__federation=federation
        ).prefetch_related(
            "adhesions__association__entite"
        ).distinct()

    elif associations.exists():
        if any(user_has_permission(request.user, ent, "membres_view_own") for ent in associations):
            qs = Personne.objects.filter(
                adhesions__association__entite__in=associations
            ).prefetch_related(
                "adhesions__association__entite"
            ).distinct()
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    q = request.GET.get("q")
    association = request.GET.get("association")
    ville = request.GET.get("ville")
    statut = request.GET.get("statut")

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q) |
            Q(numero__icontains=q)
        ).distinct()

    if association:
        qs = qs.filter(
            adhesions__association__entite__nom__icontains=association
        ).distinct()

    if ville:
        qs = qs.filter(ville__icontains=ville)

    if statut:
        qs = qs.filter(statut=statut)

    page_obj = paginate(request, qs.order_by("nom", "prenom"))

    return render(request, "personnes/liste.html", {
        "page_obj": page_obj,
        "q": q,
    })


@login_required
def personne_detail(request, pk):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = Personne.objects.none()

    federation = federations.first()

    if federation and user_has_permission(request.user, federation, "membres_view_all"):
        qs = Personne.objects.filter(
            adhesions__association__federation=federation
        ).distinct()

    elif associations_entites.exists():
        if any(user_has_permission(request.user, ent, "membres_view_own") for ent in associations_entites):
            qs = Personne.objects.filter(
                adhesions__association__entite__in=associations_entites
            ).distinct()
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    personne = get_object_or_404(qs, pk=pk)

    return render(request, "personnes/detail.html", {
        "personne": personne
    })

@login_required
def personne_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    federation = federations.first()

    can_edit = federation and user_has_permission(
        request.user,
        federation,
        "membres_edit"
    )

    can_propose = any(
        user_has_permission(request.user, entite, "membres_propose")
        for entite in associations_entites
    )

    if not can_edit and not can_propose:
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = PersonneForm(request.POST, request.FILES)

        if form.is_valid():
            email = form.cleaned_data.get("email")
            telephone = form.cleaned_data.get("telephone")
            nom = form.cleaned_data.get("nom")
            prenom = form.cleaned_data.get("prenom")
            ville = form.cleaned_data.get("ville")

            existing_qs = Personne.objects.none()

            if email:
                existing_qs = existing_qs | Personne.objects.filter(email__iexact=email)

            if telephone:
                existing_qs = existing_qs | Personne.objects.filter(telephone__iexact=telephone)

            if nom and prenom and ville:
                existing_qs = existing_qs | Personne.objects.filter(
                    nom__iexact=nom,
                    prenom__iexact=prenom,
                    ville__iexact=ville,
                )

            existing = existing_qs.distinct().first()

            if existing:
                messages.warning(
                    request,
                    "Une personne similaire existe déjà. Vérifiez avant de créer un doublon."
                )
                return redirect("personne_detail", pk=existing.pk)

            personne = form.save()

            if can_propose and not can_edit:
                association_entite = associations_entites.first()

                association = get_object_or_404(
                    Association,
                    entite=association_entite
                )

                adhesion, created = Adhesion.objects.get_or_create(
                    personne=personne,
                    association=association,
                    defaults={
                        "statut": "en_attente",
                    }
                )

                if not created:
                    messages.warning(
                        request,
                        "Cette personne est déjà liée à cette association."
                    )
                    return redirect("personnes_liste")

                admins = User.objects.filter(is_superuser=True)

                notifier_users(
                    admins,
                    f"Nouvelle demande d'adhésion proposée par {association.entite.nom} : {personne}",
                    "membre",
                    lien="/membres/a-valider/"
                )

                messages.success(
                    request,
                    "Membre proposé avec succès. Le bureau exécutif a été notifié."
                )
                return redirect("personnes_liste")

            messages.success(
                request,
                f"Personne créée avec succès. Numéro : {personne.numero}"
            )
            return redirect("personnes_liste")

    else:
        form = PersonneForm()

    return render(request, "personnes/form.html", {
        "form": form,
        "title": "Ajouter une personne"
    })


@login_required
def personne_update(request, pk):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "membres_edit"):
        return HttpResponseForbidden("Accès refusé.")

    personne = get_object_or_404(
        Personne,
        pk=pk,
        adhesions__association__federation=federation
    )

    if request.method == "POST":
        form = PersonneForm(request.POST, instance=personne)
        if form.is_valid():
            form.save()
            messages.success(request, "Personne modifiée avec succès.")
            return redirect("personne_detail", pk=personne.pk)
    else:
        form = PersonneForm(instance=personne)

    return render(request, "personnes/form.html", {
        "form": form,
        "title": "Modifier une personne"
    })

@login_required
def membres_a_valider(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "membres_edit"):
        return HttpResponseForbidden("Accès refusé.")

    adhesions = Adhesion.objects.filter(
        association__federation=federation,
        statut="en_attente"
    ).select_related("personne", "association__entite")

    return render(request, "personnes/a_valider.html", {
        "adhesions": adhesions
    })


@login_required
def adhesion_valider(request, pk):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    adhesion = get_object_or_404(
        Adhesion.objects.select_related("personne", "association__federation"),
        pk=pk,
        association__federation=federation
    )

    if not federation or not user_has_permission(request.user, federation, "membres_edit"):
        return HttpResponseForbidden("Accès refusé.")

    adhesion.statut = "valide"
    adhesion.valide_par = request.user
    adhesion.date_validation = timezone.now()
    adhesion.save()

    if not adhesion.personne.numero:
        adhesion.personne.numero = generer_numero_personne(federation)
        adhesion.personne.save()

    messages.success(request, f"Membre validé : {adhesion.personne}")
    return redirect("membres_a_valider")


@login_required
def adhesion_refuser(request, pk):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    adhesion = get_object_or_404(
        Adhesion.objects.select_related("personne", "association__federation"),
        pk=pk,
        association__federation=federation
    )

    if not federation or not user_has_permission(request.user, federation, "membres_edit"):
        return HttpResponseForbidden("Accès refusé.")

    adhesion.statut = "refuse"
    adhesion.valide_par = request.user
    adhesion.date_validation = timezone.now()
    adhesion.save()

    messages.warning(request, f"Demande refusée : {adhesion.personne}")
    return redirect("membres_a_valider")

@login_required
def cotisations_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = CotisationAssociationMensuelle.objects.none()
    contributions_deces = ContributionDecesAssociation.objects.none()
    paiements_avance = PaiementAssociation.objects.none()

    federation = federations.first()

    if federation and user_has_permission(request.user, federation, "cotisations_view_all"):
        qs = CotisationAssociationMensuelle.objects.filter(
            affiliation__federation=federation
        ).select_related(
            "affiliation__association__entite",
            "affiliation__federation",
        )

        contributions_deces = ContributionDecesAssociation.objects.filter(
            dossier__affiliation__federation=federation
        ).select_related("association__entite")

        paiements_avance = PaiementAssociation.objects.filter(
            association__federation=federation
        )

    elif associations_entites.exists():
        if any(user_has_permission(request.user, ent, "cotisations_view_own") for ent in associations_entites):
            associations_ids = Association.objects.filter(
                entite__in=associations_entites
            ).values_list("id", flat=True)

            qs = CotisationAssociationMensuelle.objects.filter(
                affiliation__association_id__in=associations_ids
            ).select_related(
                "affiliation__association__entite",
                "affiliation__federation",
            )

            contributions_deces = ContributionDecesAssociation.objects.filter(
                association_id__in=associations_ids
            ).select_related("association__entite")

            paiements_avance = PaiementAssociation.objects.filter(
                association_id__in=associations_ids
            )
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    code = request.GET.get("code")
    annee = request.GET.get("annee")
    mois = request.GET.get("mois")
    statut = request.GET.get("statut")
    association = request.GET.get("association")

    if code:
        qs = qs.filter(affiliation__association__code_paiement__icontains=code)
        contributions_deces = contributions_deces.filter(
            association__code_paiement__icontains=code
        )
        paiements_avance = paiements_avance.filter(
            association__code_paiement__icontains=code
        )

    if association:
        qs = qs.filter(affiliation__association__entite__nom__icontains=association)
        contributions_deces = contributions_deces.filter(
            association__entite__nom__icontains=association
        )
        paiements_avance = paiements_avance.filter(
            association__entite__nom__icontains=association
        )

    if annee:
        qs = qs.filter(annee=annee)

    if mois:
        qs = qs.filter(mois=mois)

    if statut:
        qs = qs.filter(statut=statut)
        contributions_deces = contributions_deces.filter(statut=statut)

    resume = defaultdict(lambda: {
        "code": "",
        "association": "",
        "mensuel_attendu": 0,
        "mensuel_paye": 0,
        "deces_attendu": 0,
        "deces_paye": 0,
        "total_attendu": 0,
        "total_paye": 0,
        "reste": 0,
        "avance": 0,
        "situation": "",
    })

    cotisations_resume = (
        qs.values(
            "affiliation__association_id",
            "affiliation__association__entite__nom",
            "affiliation__association__code_paiement",
        )
        .annotate(
            total_attendu=Sum("montant_attendu"),
            total_paye=Sum("montant_paye"),
        )
        .order_by("affiliation__association__entite__nom")
    )

    for item in cotisations_resume:
        association_id = item["affiliation__association_id"]

        resume[association_id]["code"] = item["affiliation__association__code_paiement"]
        resume[association_id]["association"] = item["affiliation__association__entite__nom"]
        resume[association_id]["mensuel_attendu"] = item["total_attendu"] or 0
        resume[association_id]["mensuel_paye"] = item["total_paye"] or 0

    deces_resume = (
        contributions_deces.values(
            "association_id",
            "association__entite__nom",
            "association__code_paiement",
        )
        .annotate(
            total_attendu=Sum("montant_attendu"),
            total_paye=Sum("montant_paye"),
        )
        .order_by("association__entite__nom")
    )

    for item in deces_resume:
        association_id = item["association_id"]

        resume[association_id]["code"] = item["association__code_paiement"]
        resume[association_id]["association"] = item["association__entite__nom"]
        resume[association_id]["deces_attendu"] = item["total_attendu"] or 0
        resume[association_id]["deces_paye"] = item["total_paye"] or 0

    avances_resume = (
        paiements_avance.values(
            "association_id",
            "association__entite__nom",
            "association__code_paiement",
        )
        .annotate(
            total_avance=Sum("montant_avance")
        )
        .order_by("association__entite__nom")
    )

    for item in avances_resume:
        association_id = item["association_id"]

        resume[association_id]["code"] = item["association__code_paiement"]
        resume[association_id]["association"] = item["association__entite__nom"]
        resume[association_id]["avance"] = item["total_avance"] or 0

    resume_associations = []

    for item in resume.values():
        item["total_attendu"] = item["mensuel_attendu"] + item["deces_attendu"]
        item["total_paye"] = item["mensuel_paye"] + item["deces_paye"]

        avance_existante = item.get("avance", 0)
        solde_net = item["total_attendu"] - item["total_paye"] - avance_existante

        if solde_net > 0:
            item["reste"] = solde_net
            item["avance"] = 0
            item["situation"] = "En retard"
        else:
            item["reste"] = 0
            item["avance"] = abs(solde_net)

            if item["avance"] > 0:
                item["situation"] = "En avance"
            else:
                item["situation"] = "À jour"

        resume_associations.append(item)

    resume_associations = sorted(
        resume_associations,
        key=lambda x: x["association"]
    )

    total_attendu_global = sum(item["total_attendu"] for item in resume_associations)
    total_paye_global = sum(item["total_paye"] for item in resume_associations)
    total_reste_global = sum(item["reste"] for item in resume_associations)
    total_avance_global = sum(item["avance"] for item in resume_associations)

    afficher_detail_mensuel = bool(code or association or annee or mois or statut)

    page_obj = paginate(request, qs.order_by("-annee", "-mois"))

    return render(request, "cotisations/liste.html", {
        "page_obj": page_obj,
        "resume_associations": resume_associations,
        "total_attendu": total_attendu_global,
        "total_paye": total_paye_global,
        "total_reste": total_reste_global,
        "total_avance": total_avance_global,
        "afficher_detail_mensuel": afficher_detail_mensuel,
    })

@login_required
def cotisation_pdf(request, pk):
    cotisation = get_object_or_404(
        CotisationAssociationMensuelle.objects.select_related(
            "affiliation__association__entite",
            "affiliation__federation",
        ),
        pk=pk,
    )

    federation = cotisation.affiliation.federation

    if not user_has_permission(request.user, federation, "cotisations_view_all"):
        return HttpResponseForbidden("Accès refusé.")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="cotisation_{cotisation.pk}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(
        Paragraph(
            f"Cotisation {cotisation.mois}/{cotisation.annee}",
            styles["Title"]
        )
    )

    elements.append(Spacer(1, 12))

    data = [
        ["Association", cotisation.affiliation.association.entite.nom],
        ["Code", cotisation.affiliation.association.code_paiement],
        ["Fédération", cotisation.affiliation.federation.nom],
        ["Montant attendu", str(cotisation.montant_attendu)],
        ["Montant payé", str(cotisation.montant_paye)],
        ["Statut", cotisation.get_statut_display()],
    ]

    table = Table(data, colWidths=[180, 300])

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(table)

    doc.build(elements)

    return response


@login_required
def cotisation_detail(request, pk):
    cotisation = get_object_or_404(
        CotisationAssociationMensuelle.objects.select_related(
            "affiliation__association__entite",
            "affiliation__federation",
        ),
        pk=pk,
    )

    can_view = (
        user_has_permission(request.user, cotisation.affiliation.federation, "cotisations_view_all")
        or user_has_permission(request.user, cotisation.affiliation.association.entite, "cotisations_view_own")
    )

    if not can_view:
        return HttpResponseForbidden("Accès refusé.")

    return render(request, "cotisations/detail.html", {"cotisation": cotisation})


@login_required
def cotisation_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_edit"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = CotisationAssociationMensuelleForm(request.POST)
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations,
            statut="active",
        ).select_related("association__entite", "federation")

        if form.is_valid():
            affiliation = form.cleaned_data["affiliation"]
            annee = form.cleaned_data["annee"]
            mois = form.cleaned_data["mois"]

            exists = CotisationAssociationMensuelle.objects.filter(
                affiliation=affiliation,
                annee=annee,
                mois=mois,
            ).exists()

            if exists:
                messages.warning(
                    request,
                    "Cette cotisation existe déjà pour cette association et cette période."
                )
                return redirect("cotisations_liste")

            cotisation = form.save(commit=False)

            if cotisation.montant_paye >= cotisation.montant_attendu:
                cotisation.statut = "paye"
            elif cotisation.montant_paye > 0:
                cotisation.statut = "partiel"
            else:
                cotisation.statut = "en_attente"

            cotisation.modifie_par = request.user
            cotisation.date_modification_paiement = timezone.now()
            cotisation.save()

            messages.success(
                request,
                "Cotisation ajoutée manuellement avec succès."
            )
            return redirect("cotisations_liste")

    else:
        form = CotisationAssociationMensuelleForm()
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations,
            statut="active",
        ).select_related("association__entite", "federation")

    return render(request, "cotisations/form.html", {
        "form": form,
        "title": "Ajouter une cotisation manuelle"
    })

@login_required
def cotisation_update(request, pk):
    cotisation = get_object_or_404(
        CotisationAssociationMensuelle.objects.select_related("affiliation__federation"),
        pk=pk,
    )

    federation = cotisation.affiliation.federation

    if not user_has_permission(request.user, federation, "cotisations_edit"):
        return HttpResponseForbidden("Accès refusé.")

    can_override = user_has_permission(
        request.user,
        federation,
        "finance_override"
    )

    # Protection finance : cotisation payée verrouillée sauf admin autorisé
    if cotisation.statut == "paye" and not can_override:
        messages.warning(
            request,
            "Cette cotisation est déjà payée et ne peut plus être modifiée."
        )
        return redirect("cotisation_detail", pk=cotisation.pk)

    federations = get_user_entites(request.user).filter(type="federation")

    if request.method == "POST":
        ancien_statut = cotisation.statut

        form = CotisationAssociationMensuelleForm(request.POST, instance=cotisation)
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations
        )

        if form.is_valid():
            cotisation = form.save(commit=False)
            cotisation.modifie_par =request.user
            cotisation.date_modification_paiement = timezone.now()
            cotisation.save()

            if cotisation.statut == "paye" and ancien_statut != "paye":
                try:
                    ok = envoyer_confirmation_paiement(cotisation)
                    if ok:
                        messages.success(request, "Cotisation confirmée et email de confirmation envoyé.")
                    else:
                        messages.warning(request, "Cotisation confirmée, mais l'association n'a pas d'email.")
                except Exception:
                    messages.warning(request, "Cotisation confirmée, mais email non envoyé.")
            else:
                messages.success(request, "Cotisation modifiée avec succès.")

            return redirect("cotisation_detail", pk=cotisation.pk)

    else:
        form = CotisationAssociationMensuelleForm(instance=cotisation)
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations
        )

    return render(request, "cotisations/form.html", {
        "form": form,
        "title": "Modifier une cotisation",
    })


@login_required
def export_cotisations_excel(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_view_all"):
        return HttpResponseForbidden("Accès refusé.")

    resume = get_resume_financier_associations(federation)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotisations"

    ws.append([
        "Code",
        "Association",
        "Mensuel dû",
        "Mensuel payé",
        "Décès dû",
        "Décès payé",
        "Total dû",
        "Total payé",
        "Reste",
        "Avance",
        "Situation",
    ])

    for item in resume:
        ws.append([
            item["code"],
            item["association"],
            item["mensuel_attendu"],
            item["mensuel_paye"],
            item["deces_attendu"],
            item["deces_paye"],
            item["total_attendu"],
            item["total_paye"],
            item["reste"],
            item["avance"],
            item["situation"],
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cotisations_resume.xlsx"'

    wb.save(response)
    return response

@login_required
def export_cotisations_pdf(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_view_all"):
        return HttpResponseForbidden("Accès refusé.")

    resume = get_resume_financier_associations(federation)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cotisations_resume.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Résumé financier des cotisations", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [[
        "Code",
        "Association",
        "Total dû",
        "Total payé",
        "Reste",
        "Avance",
        "Situation",
    ]]

    for item in resume:
        data.append([
            item["code"],
            item["association"],
            str(item["total_attendu"]),
            str(item["total_paye"]),
            str(item["reste"]),
            str(item["avance"]),
            item["situation"],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(table)
    doc.build(elements)

    return response
@login_required
def generer_cotisations(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_generate"):
        return HttpResponseForbidden("Accès refusé.")

    try:
        annee = int(request.GET.get("annee", date.today().year))
        mois = int(request.GET.get("mois", date.today().month))
    except ValueError:
        messages.error(request, "Année ou mois invalide.")
        return redirect("cotisations_liste")

    if mois < 1 or mois > 12:
        messages.error(request, "Le mois doit être compris entre 1 et 12.")
        return redirect("cotisations_liste")

    affiliations = AffiliationAssociation.objects.filter(
        federation__in=federations,
        statut="active",
    ).select_related("association", "association__entite")

    created = 0
    skipped = 0
    auto_paid = 0
    partial_paid = 0

    for aff in affiliations:
        cotisation, was_created = CotisationAssociationMensuelle.objects.get_or_create(
            affiliation=aff,
            annee=annee,
            mois=mois,
            defaults={
                "montant_attendu": aff.montant_mensuel,
                "montant_paye": 0,
                "statut": "en_attente",
            },
        )

        if not was_created:
            skipped += 1
            continue

        created += 1

        avances = PaiementAssociation.objects.filter(
            association=aff.association,
            montant_avance__gt=0
        ).order_by("date_paiement", "created_at")

        reste_a_payer = cotisation.montant_attendu - cotisation.montant_paye

        for paiement in avances:
            if reste_a_payer <= 0:
                break

            montant_utilise = min(paiement.montant_avance, reste_a_payer)

            cotisation.montant_paye += montant_utilise
            cotisation.date_paiement = paiement.date_paiement
            cotisation.mode_paiement = paiement.mode_paiement
            cotisation.modifie_par = request.user
            cotisation.date_modification_paiement = timezone.now()

            paiement.montant_avance -= montant_utilise
            paiement.montant_affecte += montant_utilise
            paiement.save(update_fields=["montant_avance", "montant_affecte"])

            reste_a_payer -= montant_utilise

        if cotisation.montant_paye >= cotisation.montant_attendu:
            cotisation.statut = "paye"
            auto_paid += 1
        elif cotisation.montant_paye > 0:
            cotisation.statut = "partiel"
            partial_paid += 1
        else:
            cotisation.statut = "en_attente"

        cotisation.save()

    messages.success(
        request,
        f"{created} cotisations générées, {skipped} déjà existantes. "
        f"{auto_paid} payées automatiquement par avance, {partial_paid} partiellement payées."
    )

    return redirect("cotisations_liste")

@login_required
def beneficiaires_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = BeneficiaireFederation.objects.none()

    federation = federations.first()

    if federation and user_has_permission(request.user, federation, "beneficiaires_view"):
        qs = BeneficiaireFederation.objects.filter(
            affiliation__federation=federation,
            personne__statut="actif"
        ).select_related(
            "personne",
            "affiliation__association__entite",
            "affiliation__federation",
        )

    elif associations_entites.exists():
        if any(user_has_permission(request.user, ent, "beneficiaires_view_own") for ent in associations_entites):
            association_ids = Association.objects.filter(
                entite__in=associations_entites
            ).values_list("id", flat=True)

            qs = BeneficiaireFederation.objects.filter(
                affiliation__association_id__in=association_ids,
                personne__statut="actif"
            ).select_related(
                "personne",
                "affiliation__association__entite",
                "affiliation__federation",
            )
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    page_obj = paginate(request, qs.order_by("-actif", "-date_activation"))

    return render(request, "beneficiaires/liste.html", {
        "page_obj": page_obj
    })


@login_required
def beneficiaire_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "beneficiaires_edit"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = BeneficiaireFederationForm(request.POST)
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Bénéficiaire ajouté avec succès.")
            return redirect("beneficiaires_liste")
    else:
        form = BeneficiaireFederationForm()
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations
        )

    return render(request, "beneficiaires/form.html", {
        "form": form,
        "title": "Ajouter un bénéficiaire"
    })

@login_required
def beneficiaire_toggle_actif(request, pk):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "beneficiaires_edit"):
        return HttpResponseForbidden("Accès refusé.")

    beneficiaire = get_object_or_404(
        BeneficiaireFederation,
        pk=pk,
        affiliation__federation=federation,
    )

    if request.method == "POST":
        if beneficiaire.actif:
            motif = request.POST.get("motif_desactivation", "").strip()

            beneficiaire.actif = False
            beneficiaire.motif_desactivation = motif or "Désactivation manuelle"
            beneficiaire.date_desactivation = timezone.now()
            beneficiaire.desactive_par = request.user

            messages.success(request, "Bénéficiaire désactivé avec succès.")

        else:
            beneficiaire.actif = True
            beneficiaire.motif_desactivation = ""
            beneficiaire.date_reactivation = timezone.now()
            beneficiaire.reactive_par = request.user

            messages.success(request, "Bénéficiaire réactivé avec succès.")

        beneficiaire.save()

    return redirect("beneficiaires_liste")

@login_required
def rapatriements_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "rapatriements_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = DossierRapatriement.objects.filter(
        affiliation__federation__in=federations
    ).select_related("personne", "affiliation__association__entite")

    page_obj = paginate(request, qs.order_by("-created_at"))
    return render(request, "rapatriements/liste.html", {"page_obj": page_obj})


@login_required
def rapatriement_detail(request, pk):
    entites = get_user_entites(request.user)
    federations = entites.filter(type="federation")
    associations_entites = entites.filter(type="association")

    dossier = get_object_or_404(
        DossierRapatriement.objects.select_related(
            "personne",
            "affiliation__association__entite",
            "affiliation__federation",
        ),
        pk=pk,
    )

    federation_dossier = dossier.affiliation.federation
    federation_user = federations.first()

    can_view = False
    can_edit = False

    # Bureau fédéral
    if federation_user and federation_user == federation_dossier:
        can_view = user_has_permission(
            request.user,
            federation_user,
            "rapatriements_view"
        )

        can_edit = user_has_permission(
            request.user,
            federation_user,
            "rapatriements_edit"
        )

    # Président association de la même fédération
    if not can_view and associations_entites.exists():
        association_user_exists = Association.objects.filter(
            entite__in=associations_entites,
            federation=federation_dossier,
        ).exists()

        if association_user_exists:
            can_view = any(
                user_has_permission(
                    request.user,
                    entite,
                    "rapatriements_view"
                )
                for entite in associations_entites
            )

    if not can_view:
        return HttpResponseForbidden("Accès refusé.")

    contributions = ContributionDecesAssociation.objects.filter(
        dossier=dossier
    ).select_related(
        "association__entite"
    ).order_by("association__entite__nom")

    total_attendu = contributions.aggregate(
        total=Sum("montant_attendu")
    )["total"] or 0

    total_paye = contributions.aggregate(
        total=Sum("montant_paye")
    )["total"] or 0

    total_reste = total_attendu - total_paye

    return render(request, "rapatriements/detail.html", {
        "dossier": dossier,
        "contributions": contributions,
        "total_attendu": total_attendu,
        "total_paye": total_paye,
        "total_reste": total_reste,
        "can_edit_rapatriement": can_edit,
    })

@login_required
def rapatriement_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "rapatriements_edit"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = DossierRapatriementForm(request.POST)
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations
        )

        if form.is_valid():
            dossier = form.save()

            # Marquer la personne comme décédée
            personne = dossier.personne
            personne.statut = "decede"
            personne.date_deces = dossier.date_deces
            personne.save()

            # Désactiver le bénéficiaire
            BeneficiaireFederation.objects.filter(
                personne=personne
            ).update(
                actif=False,
                motif_desactivation="Décès",
                date_desactivation=timezone.now(),
                desactive_par=request.user,
            )

            # Générer les contributions décès selon le coût réel
            if dossier.cout_total > 0 and not dossier.contributions_generees:
                nb = generer_contributions_deces(dossier)
                associations = Association.objects.filter(
                    federation=dossier.affiliation.federation
                    ).select_related("entite")

                emails = [
                    association.email
                    for association in associations
                    if association.email
                ]

                if emails:
                    envoyer_email_historise(
                    sujet="Nouveau dossier de rapatriement",
                    message=(
                        f"Bonjour,\n\n"
                        f"Un nouveau dossier de rapatriement a été créé.\n\n"
                        f"Défunt : {dossier.personne}\n"
                        f"Association : {dossier.affiliation.association.entite.nom}\n"
                        f"Date décès : {dossier.date_deces}\n"
                        f"Lieu décès : {dossier.lieu_deces}\n"
                        f"Pays décès : {dossier.pays_deces or '-'}\n"
                        f"Destination : {dossier.destination_rapatriement}\n"
                        f"Coût total : {dossier.cout_total} €\n"
                        f"Part par association : {dossier.montant_par_association} €\n\n"
                        f"Merci de consulter votre espace membre pour le suivi.\n\n"
                        f"Cordialement,\n"
                        f"FAG NRW e.V."
                        ),
                        destinataires=emails,
                        type_email="rapatriement",
                        utilisateur=request.user,
                        entite=dossier.affiliation.federation,
                     )
            messages.success(
                    request,
                    f"Dossier créé avec succès. {nb} contributions décès générées automatiquement."
                )
        else:
                messages.success(request, "Dossier créé avec succès.")

        return redirect("rapatriements_liste")

    else:
        form = DossierRapatriementForm()
        form.fields["affiliation"].queryset = AffiliationAssociation.objects.filter(
            federation__in=federations
        )

    return render(request, "rapatriements/form.html", {
        "form": form,
        "title": "Créer un dossier de rapatriement"
    })

@login_required
def activites_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = ActiviteCulturelle.objects.none()

    federation = federations.first()
    if federation and user_has_permission(request.user, federation, "activites_view"):
        qs = ActiviteCulturelle.objects.filter(entite__in=federations)

    elif associations_entites.exists():
        if any(user_has_permission(request.user, ent, "activites_view") for ent in associations_entites):
            federation_ids = Association.objects.filter(
                entite__in=associations_entites
            ).values_list("federation_id", flat=True)

            qs = ActiviteCulturelle.objects.filter(id__in=[])  # fallback safe
            if federation_ids:
                qs = ActiviteCulturelle.objects.filter(entite_id__in=federation_ids)
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    page_obj = paginate(request, qs.order_by("-date"))
    return render(request, "activites/liste.html", {"page_obj": page_obj})


@login_required
def activite_detail(request, pk):
    activite = get_object_or_404(
        ActiviteCulturelle.objects.select_related("entite"),
        pk=pk
    )

    can_view = (
        user_has_permission(request.user, activite.entite, "activites_view")
        or user_has_permission(request.user, activite.entite, "documents_view")
    )

    if not can_view:
        associations_entites = get_user_entites(request.user).filter(type="association")
        federation_ids = Association.objects.filter(entite__in=associations_entites).values_list("federation_id", flat=True)
        if activite.entite_id not in federation_ids:
            return HttpResponseForbidden("Accès refusé.")

    return render(request, "activites/detail.html", {"activite": activite})


@login_required
def activite_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "activites_create"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = ActiviteCulturelleForm(request.POST)
        form.fields["entite"].queryset = federations

        if form.is_valid():
            activite = form.save(commit=False)

            langue = form.cleaned_data["langue_publication"]

            if langue == "fr":
                activite.titre_fr = activite.titre
                activite.description_fr = activite.description
            elif langue == "de":
                activite.titre_de = activite.titre
                activite.description_de = activite.description
            elif langue == "en":
                activite.titre_en = activite.titre
                activite.description_en = activite.description

            activite.save()

            destinataires = list(
                Association.objects.filter(
                    federation__in=federations
                ).exclude(email="").values_list("email", flat=True)
            )

            try:
                if destinataires:
                    envoyer_notification_activite(activite, destinataires)
                    messages.success(request, "Activité créée avec succès et notification envoyée.")
                else:
                    messages.success(request, "Activité créée avec succès.")
            except Exception:
                messages.warning(request, "Activité créée, mais notification email non envoyée.")

            return redirect("activites_liste")
    else:
        form = ActiviteCulturelleForm()
        form.fields["entite"].queryset = federations

    return render(request, "activites/form.html", {
        "form": form,
        "title": "Créer une activité",
    })

@login_required
def transactions_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "transactions_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = Transaction.objects.filter(entite__in=federations).select_related("entite").order_by("-date")
    page_obj = paginate(request, qs)
    return render(request, "transactions/liste.html", {"page_obj": page_obj})

@login_required
def transaction_create(request):
    entites = get_user_entites(request.user)
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        form.fields['entite'].queryset = entites
        if form.is_valid():
            form.save()
            messages.success(request, 'Transaction créée avec succès.')
            return redirect('transactions_liste')
    else:
        form = TransactionForm()
        form.fields['entite'].queryset = entites
    return render(request, 'transactions/form.html', {'form': form, 'title': 'Ajouter une transaction'})


@login_required
def documents_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = Document.objects.none()
    federation = federations.first()

    # Bureau fédéral : voit les documents de la fédération
    # + les documents envoyés par toutes ses associations
    if federation and user_has_permission(request.user, federation, "documents_view"):
        associations_entites_ids = Association.objects.filter(
            federation=federation
        ).values_list("entite_id", flat=True)

        qs = Document.objects.filter(
            Q(entite=federation) |
            Q(entite_id__in=associations_entites_ids)
        ).select_related("entite").distinct()

    # Association : voit ses propres documents
    # + les documents fédéraux destinés à toutes les associations
    # + les documents publics
    elif associations_entites.exists():
        if any(user_has_permission(request.user, ent, "documents_view") for ent in associations_entites):
            federation_ids = Association.objects.filter(
                entite__in=associations_entites
            ).values_list("federation_id", flat=True)

            qs = Document.objects.filter(
                Q(entite__in=associations_entites) |
                Q(entite_id__in=federation_ids, visibilite__in=["associations"])
            ).select_related("entite").distinct()
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    page_obj = paginate(request, qs.order_by("-created_at"))

    return render(request, "documents/liste.html", {
        "page_obj": page_obj
    })

@login_required
def document_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    allowed_entites = Entite.objects.none()
    federation = federations.first()

    if federation and user_has_permission(request.user, federation, "documents_upload"):
        allowed_entites = federations

    elif associations_entites.exists():
        if any(user_has_permission(request.user, a, "documents_upload") for a in associations_entites):
            allowed_entites = associations_entites
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = DocumentForm(request.POST, request.FILES,User=request.user,is_federation=bool(federation))
        form.fields["entite"].queryset = allowed_entites

        if form.is_valid():
            document = form.save(commit=False)
            document.soumis_par = request.user
            document.statut = "nouveau"
            document.save()

            if document.entite.type == "association":
                association = Association.objects.filter(
                    entite=document.entite
                ).select_related("federation").first()

                if association:
                    Notification.objects.create(
                        entite=association.federation,
                        titre="Nouveau document reçu",
                        message=f"{association.entite.nom} a envoyé un document : {document.titre}",
                        type_notification="document",
                        document=document,
                    )

                    admins = User.objects.filter(is_superuser=True)
                    notifier_users(
                        admins,
                        f"Nouveau document soumis par {association.entite.nom} : {document.titre}",
                        "document",
                        lien="/documents/"
                    )

            messages.success(request, "Document envoyé avec succès. Le bureau exécutif a été notifié.")
            return redirect("documents_liste")

    else:
        form = DocumentForm(user=request.user,is_federation=bool(federation))
        form.fields["entite"].queryset = allowed_entites

    return render(request, "documents/form.html", {
        "form": form,
        "title": "Ajouter un document"
    })

@login_required
def document_update_statut(request, pk, statut):
    statuts_autorises = ["nouveau", "en_cours", "traite", "rejete"]

    if statut not in statuts_autorises:
        return HttpResponseForbidden("Statut invalide.")

    document = get_object_or_404(Document, pk=pk)

    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "documents_traiter"):
        return HttpResponseForbidden("Accès refusé.")

    document.statut = statut
    document.traite_par = request.user
    document.date_traitement = timezone.now()
    document.save()

    if document.soumis_par:
        notifier_users(
            [document.soumis_par],
            f"Votre document '{document.titre}' est maintenant : {document.get_statut_display()}",
            "document",
            lien="/documents/"
        )

    messages.success(request, "Statut du document mis à jour.")
    return redirect("documents_liste")


@login_required
def rapports(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "rapports_view"):
        return HttpResponseForbidden("Accès refusé.")

    resume_financier = get_resume_financier_associations(federation)

    total_du = sum(item["total_attendu"] for item in resume_financier)
    total_paye = sum(item["total_paye"] for item in resume_financier)
    total_reste = sum(item["reste"] for item in resume_financier)
    total_avance = sum(item["avance"] for item in resume_financier)

    associations_en_retard = [
        item for item in resume_financier if item["reste"] > 0
    ]

    associations_en_avance = [
        item for item in resume_financier if item["avance"] > 0
    ]

    cotisations_par_annee = (
        CotisationAssociationMensuelle.objects
        .filter(affiliation__federation__in=federations)
        .values("annee")
        .annotate(
            total_paye=Sum("montant_paye"),
            total_attendu=Sum("montant_attendu")
        )
        .order_by("annee")
    )

    cotisations_deces_global = ContributionDecesAssociation.objects.filter(
        dossier__affiliation__federation__in=federations
    ).aggregate(
        total_attendu=Sum("montant_attendu"),
        total_paye=Sum("montant_paye"),
    )

    deces_total_attendu = cotisations_deces_global["total_attendu"] or 0
    deces_total_paye = cotisations_deces_global["total_paye"] or 0
    deces_reste = deces_total_attendu - deces_total_paye

    beneficiaires_par_association = (
        BeneficiaireFederation.objects
        .filter(
            affiliation__federation__in=federations,
            actif=True,
            personne__statut="actif",
        )
        .values("affiliation__association__entite__nom")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    rapatriements_par_statut = (
        DossierRapatriement.objects
        .filter(affiliation__federation__in=federations)
        .values("statut")
        .annotate(total=Count("id"))
        .order_by("statut")
    )

    cotisations_en_retard = associations_en_retard[:10]

    return render(request, "rapports/liste.html", {
        "resume_financier": resume_financier,
        "total_du": total_du,
        "total_paye": total_paye,
        "total_reste": total_reste,
        "total_avance": total_avance,
        "associations_en_retard": associations_en_retard,
        "associations_en_avance": associations_en_avance,

        "cotisations_par_annee": cotisations_par_annee,
        "beneficiaires_par_association": beneficiaires_par_association,
        "rapatriements_par_statut": rapatriements_par_statut,
        "cotisations_en_retard": cotisations_en_retard,

        "deces_total_attendu": deces_total_attendu,
        "deces_total_paye": deces_total_paye,
        "deces_reste": deces_reste,
    })

@login_required
def envoyer_rappel_cotisation_view(request, pk):
    cotisation = get_object_or_404(
        CotisationAssociationMensuelle.objects.select_related(
            "affiliation__federation",
            "affiliation__association__entite",
        ),
        pk=pk,
    )

    if not user_has_permission(
        request.user,
        cotisation.affiliation.federation,
        "notifications_send"
    ):
        return HttpResponseForbidden("Accès refusé.")

    # empêcher rappel si déjà payé
    if cotisation.statut == "paye":
        messages.info(
            request,
            "Cette cotisation est déjà payée."
        )
        return redirect("cotisation_detail", pk=cotisation.pk)

    ok = envoyer_email_rappel_cotisation(cotisation)

    if ok:
        messages.success(
            request,
            "Rappel envoyé avec succès."
        )
    else:
        messages.warning(
            request,
            "Impossible d'envoyer le rappel : email de l'association manquant."
        )

    return redirect("cotisation_detail", pk=cotisation.pk)

# fonction pour rappel de deces - 17.05.2026

@login_required
def envoyer_rappel_deces_view(request, pk):
    contribution = get_object_or_404(
        ContributionDecesAssociation.objects.select_related(
            "association__entite",
            "dossier__personne",
            "dossier__affiliation__federation",
        ),
        pk=pk,
    )

    federation = contribution.dossier.affiliation.federation

    if not federation or not user_has_permission(
        request.user,
        federation,
        "envoyer_rappel_deces_view"
    ):
        return HttpResponseForbidden("Accès refusé.")

    association = contribution.association

    if not association.email:
        messages.warning(
            request,
            "Cette association ne possède pas d’adresse email."
        )
        return redirect("contributions_deces_liste")

    sujet = (
        f"Rappel cotisation décès - "
        f"{contribution.dossier.personne}"
    )

    message = (
        f"Bonjour,\n\n"
        f"Nous vous rappelons que la contribution décès suivante "
        f"reste à régler.\n\n"
        f"Défunt : {contribution.dossier.personne}\n"
        f"Association : {association.entite.nom}\n"
        f"Montant attendu : {contribution.montant_attendu} €\n"
        f"Montant payé : {contribution.montant_paye} €\n"
        f"Reste à payer : "
        f"{contribution.montant_attendu - contribution.montant_paye} €\n\n"
        f"Merci d’effectuer le paiement rapidement.\n\n"
        f"Cordialement,\n"
        f"FAG NRW e.V."
    )

    ok = envoyer_email_historise(
        sujet=sujet,
        message=message,
        destinataires=[association.email],
        type_email="rappel_deces",
        utilisateur=request.user,
        entite=federation,
    )

    if ok:
        messages.success(request, "Rappel envoyé avec succès.")
    else:
        messages.error(request, "Erreur lors de l’envoi du rappel.")

    return redirect("contributions_deces_liste")


@login_required
def emails_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "emails_view"):
        return HttpResponseForbidden("Accès refusé.")

    emails = HistoriqueEmail.objects.filter(
        Q(cotisation__affiliation__federation__in=federations) |
        Q(activite__entite__in=federations)
    ).order_by("-created_at").distinct()

    page_obj = paginate(request, emails, per_page=20)

    return render(request, "emails/liste.html", {
        "page_obj": page_obj
    })

@login_required
def informations_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = Information.objects.none()

    federation = federations.first()

    # Bureau fédéral : voit toutes les informations, même les brouillons
    if federation and user_has_permission(request.user, federation, "informations_view"):
        qs = Information.objects.filter(
            entite=federation
        )

    # Association : voit seulement les informations publiées
    elif associations_entites.exists():
        if any(user_has_permission(request.user, a, "informations_view") for a in associations_entites):
            federation_ids = Association.objects.filter(
                entite__in=associations_entites
            ).values_list("federation_id", flat=True)

            qs = Information.objects.filter(
                entite_id__in=federation_ids,
                statut="publie"
            )
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    page_obj = paginate(request, qs.order_by("-created_at"))

    return render(request, "informations/liste.html", {
        "page_obj": page_obj
    })

@login_required
def information_detail(request, pk):
    information = get_object_or_404(
        Information.objects.select_related("entite", "auteur"),
        pk=pk,
        statut="publie"
    )

    can_view = (
        user_has_permission(request.user, information.entite, "informations_view")
    )

    if not can_view:
        associations_entites = get_user_entites(request.user).filter(type="association")
        federation_ids = Association.objects.filter(
            entite__in=associations_entites
        ).values_list("federation_id", flat=True)

        if information.entite_id not in federation_ids:
            return HttpResponseForbidden("Accès refusé.")

    return render(request, "informations/detail.html", {"information": information})

@login_required
def information_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "informations_publish"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data["entite"] = federation.pk

        form = InformationForm(post_data, request.FILES)
        form.fields["entite"].queryset = federations

        if form.is_valid():
            info = form.save(commit=False)
            info.auteur = request.user
            info.entite = federation

            # Fallback : si ancien champ vide, on copie le français
            if not info.titre:
                info.titre = info.titre_fr or info.titre_de or info.titre_en

            if not info.contenu:
                info.contenu = info.contenu_fr or info.contenu_de or info.contenu_en

            if info.statut == "publie" and not info.date_publication:
                info.date_publication = timezone.now()

            info.save()
            messages.success(request, "Information publiée avec succès.")
            return redirect("informations_liste")

    else:
        form = InformationForm()
        form.fields["entite"].queryset = federations
        form.fields["entite"].initial = federation

    return render(request, "informations/form.html", {
        "form": form,
        "title": "Publier une information",
    })
#la partie publique 25.04.2026

def public_home(request):
    federation = Entite.objects.filter(type="federation").first()

    activites = ActiviteCulturelle.objects.filter(est_public=True).order_by("date")[:3]
    infos = Information.objects.filter(est_public=True, statut="publie").order_by("-created_at")[:3]

    return render(request, "public/home.html", {
        "federation": federation,
        "activites": activites,
        "infos": infos,
    })

def public_about(request):
    federation = Entite.objects.filter(type="federation").first()

    bureau = None
    membres_bureau = BureauMembre.objects.none()

    if federation:
        today = date.today()

        bureau = Bureau.objects.filter(
            entite=federation
        ).filter(
            Q(date_fin__isnull=True) | Q(date_fin__gte=today)
        ).order_by("-date_debut").first()

        if bureau:
            membres_bureau = BureauMembre.objects.filter(
                bureau=bureau,
                adhesion__statut="valide"
            ).select_related(
                "poste",
                "adhesion__personne",
                "adhesion__association__entite"
            ).order_by("poste__ordre")

    return render(request, "public/about.html", {
        "bureau": bureau,
        "membres_bureau": membres_bureau,
    })

def public_associations(request):
    associations = Association.objects.all()
    return render(request, "public/associations.html", {
        "associations": associations
    })


def public_activites(request):
    activites = ActiviteCulturelle.objects.filter(est_public=True)
    return render(request, "public/activites.html", {
        "activites": activites
    })


def public_informations(request):
    infos = Information.objects.filter(
        est_public=True,
        statut="publie"
    ).order_by("-created_at")

    return render(request, "public/informations.html", {
        "infos": infos
    })

def public_contact(request):
    if request.method == "POST":
        form = ContactPublicForm(request.POST)

        if form.is_valid():
            nom = form.cleaned_data["nom"]
            email = form.cleaned_data["email"]
            sujet = form.cleaned_data["sujet"]
            message = form.cleaned_data["message"]

            MessageContact.objects.create(
                nom=nom,
                email=email,
                sujet=sujet,
                message=message,
            )

            send_mail(
                f"[Contact public] {sujet}",
                f"Nom : {nom}\nEmail : {email}\n\nMessage :\n{message}",
                email,
                ["aliou2014barry@gmail.com"],
                fail_silently=True,
            )

            messages.success(request, "Votre message a été envoyé avec succès.")
            return redirect("public_contact")
    else:
        form = ContactPublicForm()

    return render(request, "public/contact.html", {
        "form": form
    })

@login_required
def notifications_liste(request):
    notifications = Notification.objects.filter(
        destinataire=request.user
    ).order_by("-created_at")

    notifications.filter(lu=False).update(lu=True)

    return render(request, "notifications/liste.html", {
        "notifications": notifications
    })

@login_required
def contributions_deces_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    qs = ContributionDecesAssociation.objects.none()

    federation = federations.first()

    if federation and user_has_permission(request.user, federation, "cotisations_view_all"):
        qs = ContributionDecesAssociation.objects.filter(
            dossier__affiliation__federation=federation
        ).select_related(
            "association__entite",
            "dossier__personne",
            "dossier__affiliation__federation",
        )

    elif associations_entites.exists():
        if any(user_has_permission(request.user, ent, "cotisations_view_own") for ent in associations_entites):
            association_ids = Association.objects.filter(
                entite__in=associations_entites
            ).values_list("id", flat=True)

            qs = ContributionDecesAssociation.objects.filter(
                association_id__in=association_ids
            ).select_related(
                "association__entite",
                "dossier__personne",
                "dossier__affiliation__federation",
            )
        else:
            return HttpResponseForbidden("Accès refusé.")
    else:
        return HttpResponseForbidden("Accès refusé.")

    statut = request.GET.get("statut")
    association = request.GET.get("association")

    if statut:
        qs = qs.filter(statut=statut)

    if association:
        qs = qs.filter(association__entite__nom__icontains=association)

    totals = qs.aggregate(
        attendu=Sum("montant_attendu"),
        paye=Sum("montant_paye"),
    )

    page_obj = paginate(request, qs.order_by("-created_at"))

    return render(request, "cotisations_deces/liste.html", {
        "page_obj": page_obj,
        "total_attendu": totals["attendu"] or 0,
        "total_paye": totals["paye"] or 0,
    })


@login_required
def contribution_deces_update(request, pk):
    contribution = get_object_or_404(
        ContributionDecesAssociation.objects.select_related(
            "dossier__affiliation__federation",
            "association__entite",
        ),
        pk=pk,
    )

    federation = contribution.dossier.affiliation.federation

    if not user_has_permission(request.user, federation, "cotisations_edit"):
        return HttpResponseForbidden("Accès refusé.")

    can_override = user_has_permission(
        request.user,
        federation,
        "finance_override"
    )

    # Protection financière : payé = verrouillé sauf admin autorisé
    if contribution.statut == "paye" and not can_override:
        messages.warning(
            request,
            "Cette cotisation décès est déjà payée et ne peut plus être modifiée."
        )
        return redirect("contributions_deces_liste")

    if request.method == "POST":
        form = ContributionDecesAssociationForm(
            request.POST,
            instance=contribution
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Cotisation décès modifiée avec succès.")
            return redirect("contributions_deces_liste")
    else:
        form = ContributionDecesAssociationForm(instance=contribution)

    return render(request, "cotisations_deces/form.html", {
        "form": form,
        "contribution": contribution,
        "title": "Modifier une cotisation décès",
    })

@login_required
def adhesion_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "membres_edit"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = AdhesionForm(request.POST)

        associations = Association.objects.filter(federation=federation)
        form.fields["association"].queryset = associations

        if form.is_valid():
            adhesion = form.save()

            personne = adhesion.personne
            if not personne.numero:
                personne.numero = generer_numero_personne(federation)
                personne.save()

            messages.success(
                request,
                f"Adhésion créée avec succès. Numéro membre : {personne.numero}"
            )
            return redirect("personne_detail", pk=personne.pk)
    else:
        form = AdhesionForm()
        form.fields["association"].queryset = Association.objects.filter(
            federation=federation
        )

    return render(request, "adhesions/form.html", {
        "form": form,
        "title": "Ajouter une adhésion",
    })

@login_required
def inviter_utilisateur(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "users_manage"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = InvitationUtilisateurForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data["email"]
            prenom = form.cleaned_data["prenom"]
            nom = form.cleaned_data["nom"]

            user, created = User.objects.get_or_create(
                username=email,
                defaults={
                    "email": email,
                    "first_name": prenom,
                    "last_name": nom,
                    "is_active": True,
                }
            )

            if created:
                user.set_unusable_password()
                user.save()

            reset_form = PasswordResetForm({"email": email})

            if reset_form.is_valid():
                reset_form.save(
                    request=request,
                    use_https=request.is_secure(),
                    from_email=None,
                    email_template_name="registration/password_reset_email.html",
                    subject_template_name="registration/password_reset_subject.txt",
                )

            messages.success(request, "Invitation envoyée avec succès.")
            return redirect("dashboard")
    else:
        form = InvitationUtilisateurForm()

    return render(request, "utilisateurs/inviter.html", {
        "form": form,
        "title": "Inviter un utilisateur",
    })

#  fonction pour la traduction automatique (traduit a travers deepl)

def traduire_texte(text, target_lang):
    if not text:
        return ""

    translator = deepl.Translator(settings.DEEPL_AUTH_KEY)
    result = translator.translate_text(text, target_lang=target_lang)
    return result.text


@login_required
def bureau_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    associations_entites = get_user_entites(request.user).filter(type="association")

    federation = federations.first()

    # Si l'utilisateur est une association, retrouver sa fédération
    if not federation and associations_entites.exists():
        association = Association.objects.filter(
            entite__in=associations_entites
        ).select_related("federation").first()

        if association:
            federation = association.federation

    if not federation:
        return HttpResponseForbidden("Accès refusé.")

    can_view = (
        user_has_permission(request.user, federation, "bureau_view")
        or any(user_has_permission(request.user, entite, "bureau_view") for entite in associations_entites)
    )

    if not can_view:
        return HttpResponseForbidden("Accès refusé.")

    today = date.today()

    bureau = Bureau.objects.filter(
        entite=federation
    ).filter(
        Q(date_fin__isnull=True) | Q(date_fin__gte=today)
    ).order_by("-date_debut").first()

    membres = BureauMembre.objects.none()

    if bureau:
        membres = BureauMembre.objects.filter(
            bureau=bureau,
            adhesion__statut="valide"
        ).select_related(
            "poste",
            "adhesion__personne",
            "adhesion__association__entite"
        ).order_by("poste__ordre")

    president = membres.filter(poste__nom__icontains="président").first()

    anciens_bureaux = Bureau.objects.filter(
        entite=federation,
        date_fin__lt=today
    ).order_by("-date_debut")

    return render(request, "bureau/liste.html", {
        "bureau": bureau,
        "membres": membres,
        "president": president,
        "anciens_bureaux": anciens_bureaux,
        "title": "Bureau",
    })


@login_required
def bureau_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "users_manage"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = BureauForm(request.POST)
        form.fields["entite"].queryset = federations

        if form.is_valid():
            form.save()
            messages.success(request, "Bureau créé avec succès.")
            return redirect("bureau_liste")
    else:
        form = BureauForm(initial={"entite": federation})
        form.fields["entite"].queryset = federations

    return render(request, "bureau/form.html", {
        "form": form,
        "title": "Créer un bureau",
    })


@login_required
def bureau_membre_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "users_manage"):
        return HttpResponseForbidden("Accès refusé.")

    bureau = Bureau.objects.filter(
        entite=federation
    ).filter(
        Q(date_fin__isnull=True) | Q(date_fin__gte=date.today())
    ).order_by("-date_debut").first()

    if not bureau:
        messages.error(request, "Aucun bureau actif trouvé.")
        return redirect("bureau_liste")

    if request.method == "POST":
        form = BureauMembreForm(request.POST, bureau=bureau)

        if form.is_valid():
            membre = form.save(commit=False)
            membre.bureau = bureau
            membre.save()

            messages.success(request, "Membre ajouté au bureau.")
            return redirect("bureau_liste")
    else:
        form = BureauMembreForm(bureau=bureau)
        form.fields["bureau"].initial = bureau

    return render(request, "bureau/membre_form.html", {
        "form": form,
        "bureau": bureau,
        "title": "Ajouter un membre au bureau",
    })


@login_required
def postes_liste(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "users_manage"):
        return HttpResponseForbidden("Accès refusé.")

    postes = Poste.objects.all().order_by("ordre", "nom")

    return render(request, "bureau/postes_liste.html", {
        "postes": postes,
        "title": "Postes du bureau",
    })




@login_required
def poste_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "users_manage"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = PosteForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Poste créé avec succès.")
            return redirect("postes_liste")
    else:
        form = PosteForm()

    return render(request, "bureau/form.html", {
        "form": form,
        "title": "Créer un poste",
    })

@login_required
def export_personnes_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Personnes"

    headers = [
        "Numero",
        "Nom",
        "Prénom",
        "Email",
        "Téléphone",
        "Ville",
    ]

    ws.append(headers)

    personnes = Personne.objects.all().order_by("nom")

    for personne in personnes:
        ws.append([
            personne.numero,
            personne.nom,
            personne.prenom,
            personne.email,
            personne.telephone,
            personne.ville,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="personnes.xlsx"'
    )

    wb.save(response)

    return response

@login_required
def export_contributions_deces_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotisations décès"

    headers = [
        "Association",
        "Dossier",
        "Montant attendu",
        "Montant payé",
        "Statut",
        "Date paiement",
    ]

    ws.append(headers)

    contributions = ContributionDecesAssociation.objects.select_related(
        "association__entite",
        "dossier",
    ).all()

    for item in contributions:
        ws.append([
            item.association.entite.nom,
            item.dossier.id,
            float(item.montant_attendu),
            float(item.montant_paye),
            item.get_statut_display(),
            item.date_paiement.strftime("%d/%m/%Y") if item.date_paiement else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="cotisations_deces.xlsx"'
    )

    wb.save(response)

    return response

@login_required
def export_contributions_deces_pdf(request):
    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="cotisations_deces.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        "Cotisations décès",
        styles["Heading1"],
    )

    elements.append(title)
    elements.append(Spacer(1, 20))

    data = [[
        "Association",
        "Montant attendu",
        "Montant payé",
        "Statut",
    ]]

    contributions = ContributionDecesAssociation.objects.select_related(
        "association__entite"
    ).all()

    for item in contributions:
        data.append([
            item.association.entite.nom,
            str(item.montant_attendu),
            str(item.montant_paye),
            item.get_statut_display(),
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    elements.append(table)

    doc.build(elements)

    return response

@login_required
def paiement_association_create(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_edit"):
        return HttpResponseForbidden("Accès refusé.")

    if request.method == "POST":
        form = PaiementAssociationForm(
            request.POST,
            federation=federation
        )

        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.enregistre_par = request.user
            paiement.save()

            association = paiement.association
            montant_restant = paiement.montant

            # 1. Payer d'abord les cotisations mensuelles anciennes
            cotisations = CotisationAssociationMensuelle.objects.filter(
                affiliation__association=association
            ).exclude(
                statut="paye"
            ).order_by("annee", "mois")

            for cotisation in cotisations:
                if montant_restant <= 0:
                    break

                dette = cotisation.montant_attendu - cotisation.montant_paye

                if dette <= 0:
                    continue

                montant_applique = min(montant_restant, dette)

                cotisation.montant_paye += montant_applique
                cotisation.date_paiement = paiement.date_paiement
                cotisation.mode_paiement = paiement.mode_paiement
                cotisation.modifie_par = request.user
                cotisation.date_modification_paiement = timezone.now()

                if cotisation.montant_paye >= cotisation.montant_attendu:
                    cotisation.statut = "paye"
                else:
                    cotisation.statut = "partiel"

                cotisation.save()
                montant_restant -= montant_applique

            # 2. Payer ensuite les cotisations décès / rapatriement
            contributions = ContributionDecesAssociation.objects.filter(
                association=association
            ).exclude(
                statut="paye"
            ).order_by("created_at")

            for contribution in contributions:
                if montant_restant <= 0:
                    break

                dette = contribution.montant_attendu - contribution.montant_paye

                if dette <= 0:
                    continue

                montant_applique = min(montant_restant, dette)

                contribution.montant_paye += montant_applique
                contribution.date_paiement = paiement.date_paiement

                if contribution.montant_paye >= contribution.montant_attendu:
                    contribution.statut = "paye"
                else:
                    contribution.statut = "partiel"

                contribution.save()
                montant_restant -= montant_applique

            # 3. Le reste devient une avance
            paiement.montant_affecte = paiement.montant - montant_restant
            paiement.montant_avance = montant_restant
            paiement.save(update_fields=["montant_affecte", "montant_avance"])

            # 4. Email de confirmation
            president = association.get_president()

            destinataires = []

            if association.email:
                destinataires.append(association.email)

            if president and president.email and president.email not in destinataires:
                destinataires.append(president.email)

            if destinataires:
                message_email = (
                    f"Bonjour,\n\n"
                    f"Nous confirmons la réception de votre paiement.\n\n"
                    f"Association : {association.entite.nom}\n"
                    f"Identifiant : {association.code_paiement}\n"
                    f"Montant reçu : {paiement.montant} €\n"
                    f"Montant affecté : {paiement.montant_affecte} €\n"
                    f"Avance disponible : {paiement.montant_avance} €\n"
                    f"Date : {paiement.date_paiement}\n"
                    f"Mode : {paiement.get_mode_paiement_display()}\n\n"
                    f"Cordialement,\n"
                    f"Bureau exécutif"
                )

                send_mail(
                    subject="Confirmation de paiement",
                    message=message_email,
                    from_email=None,
                    recipient_list=destinataires,
                    fail_silently=True,
                )

            # 5. Lien WhatsApp
            if president and president.telephone:
                telephone = (
                    president.telephone
                    .replace("+", "")
                    .replace(" ", "")
                    .replace("-", "")
                )

                whatsapp_message = quote(
                    f"Bonjour {president.prenom}, "
                    f"nous confirmons la réception du paiement de {paiement.montant} € "
                    f"pour {association.entite.nom}. "
                    f"Montant affecté : {paiement.montant_affecte} €. "
                    f"Avance disponible : {paiement.montant_avance} €. "
                    f"Identifiant : {association.code_paiement}."
                )

                whatsapp_url = f"https://wa.me/{telephone}?text={whatsapp_message}"

                messages.info(
                    request,
                    mark_safe(
                        f'<a href="{whatsapp_url}" target="_blank" class="alert-link">'
                        f'Envoyer aussi la confirmation WhatsApp au président'
                        f'</a>'
                    )
                )

            if montant_restant > 0:
                messages.warning(
                    request,
                    f"Paiement enregistré. Avance disponible : {montant_restant}."
                )
            else:
                messages.success(
                    request,
                    "Paiement enregistré et réparti automatiquement."
                )

            return redirect("cotisations_liste")

    else:
        form = PaiementAssociationForm(
            federation=federation
        )

    return render(request, "paiements_associations/form.html", {
        "form": form,
        "title": "Encaisser un paiement association",
    })

def arrondir_montant(montant, multiple=Decimal("5")):
    return (
        (montant/multiple)
        .quantize(Decimal("1"),rounding=ROUND_UP)
        * multiple
    )

@transaction.atomic
def generer_contributions_deces(dossier):
    federation = dossier.affiliation.federation

    associations = Association.objects.filter(
        federation=federation,
        affiliations__statut="active"
    ).distinct()

    total_associations = associations.count()

    if total_associations == 0:
        return 0

    montant_brut = Decimal(dossier.cout_total) / Decimal(total_associations)
    montant_part = arrondir_montant(montant_brut)

    dossier.montant_par_association = montant_part
    dossier.contributions_generees = True
    dossier.save(update_fields=[
        "montant_par_association",
        "contributions_generees",
    ])

    created = 0

    for association in associations:
        contribution, was_created = ContributionDecesAssociation.objects.get_or_create(
            dossier=dossier,
            association=association,
            defaults={
                "montant_attendu": montant_part,
                "montant_paye": 0,
                "statut": "en_attente",
            }
        )

        if was_created:
            created += 1

        reste_a_payer = contribution.montant_attendu - contribution.montant_paye

        avances = PaiementAssociation.objects.filter(
            association=association,
            montant_avance__gt=0
        ).order_by("date_paiement", "created_at")

        for paiement in avances:
            if reste_a_payer <= 0:
                break

            montant_utilise = min(paiement.montant_avance, reste_a_payer)

            contribution.montant_paye += montant_utilise
            contribution.date_paiement = paiement.date_paiement

            paiement.montant_avance -= montant_utilise
            paiement.montant_affecte += montant_utilise
            paiement.save(update_fields=["montant_avance", "montant_affecte"])

            reste_a_payer -= montant_utilise

        if contribution.montant_paye >= contribution.montant_attendu:
            contribution.statut = "paye"
        elif contribution.montant_paye > 0:
            contribution.statut = "partiel"
        else:
            contribution.statut = "impaye"

        contribution.save()

    total_collecte = ContributionDecesAssociation.objects.filter(
        dossier=dossier
    ).aggregate(total=Sum("montant_paye"))["total"] or Decimal("0")

    total_attendu = ContributionDecesAssociation.objects.filter(
        dossier=dossier
    ).aggregate(total=Sum("montant_attendu"))["total"] or Decimal("0")

    dossier.montant_collecte = total_collecte
    dossier.reste_a_collecter = total_attendu - total_collecte
    dossier.save(update_fields=[
        "montant_collecte",
        "reste_a_collecter",
    ])

    return created
def export_vide(request, message="Aucune donnée à exporter."):
    messages.warning(request, message)
    return redirect(request.META.get("HTTP_REFERER", "cotisations_liste"))


def filtrer_cotisations_mensuelles(qs, request):
    association = request.GET.get("association")
    code = request.GET.get("code")
    annee = request.GET.get("annee")
    mois = request.GET.get("mois")
    statut = request.GET.get("statut")
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    if association:
        qs = qs.filter(affiliation__association__entite__nom__icontains=association)

    if code:
        qs = qs.filter(affiliation__association__code_paiement__icontains=code)

    if annee:
        qs = qs.filter(annee=annee)

    if mois:
        qs = qs.filter(mois=mois)

    if statut:
        qs = qs.filter(statut=statut)

    debut = parse_date(date_debut) if date_debut else None
    fin = parse_date(date_fin) if date_fin else None

    if debut and fin:
        qs = qs.filter(
            Q(annee__gt=debut.year) |
            Q(annee=debut.year, mois__gte=debut.month)
        ).filter(
            Q(annee__lt=fin.year) |
            Q(annee=fin.year, mois__lte=fin.month)
        )

    elif debut:
        qs = qs.filter(
            Q(annee__gt=debut.year) |
            Q(annee=debut.year, mois__gte=debut.month)
        )

    elif fin:
        qs = qs.filter(
            Q(annee__lt=fin.year) |
            Q(annee=fin.year, mois__lte=fin.month)
        )

    return qs


def filtrer_contributions_deces(qs, request):
    association = request.GET.get("association")
    statut = request.GET.get("statut")
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    if association:
        qs = qs.filter(association__entite__nom__icontains=association)

    if statut:
        qs = qs.filter(statut=statut)

    if date_debut:
        qs = qs.filter(date_paiement__gte=date_debut)

    if date_fin:
        qs = qs.filter(date_paiement__lte=date_fin)

    return qs


def filtrer_membres(qs, request):
    association = request.GET.get("association")
    statut = request.GET.get("statut")
    ville = request.GET.get("ville")
    q = request.GET.get("q")

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q) |
            Q(numero__icontains=q)
        )

    if association:
        qs = qs.filter(adhesions__association__entite__nom__icontains=association)

    if ville:
        qs = qs.filter(ville__icontains=ville)

    if statut:
        qs = qs.filter(statut=statut)

    return qs.distinct()


@login_required
def export_cotisations_detail_excel(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_export"):
        return HttpResponseForbidden("Accès refusé.")

    qs = CotisationAssociationMensuelle.objects.filter(
        affiliation__federation=federation
    ).select_related(
        "affiliation__association__entite",
        "modifie_par",
    )

    qs = filtrer_cotisations_mensuelles(qs, request)

    if not qs.exists():
        return export_vide(request, "Aucune cotisation trouvée pour cet export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Détail mensuel"

    ws.append([
        "Association",
        "Identifiant",
        "Mois",
        "Année",
        "Attendu",
        "Payé",
        "Reste",
        "Statut",
        "Date paiement",
        "Mode paiement",
        "Modifié par",
        "Date modification",
    ])

    for c in qs.order_by("-annee", "-mois"):
        reste = c.montant_attendu - c.montant_paye

        ws.append([
            c.affiliation.association.entite.nom,
            c.affiliation.association.code_paiement,
            c.mois,
            c.annee,
            c.montant_attendu,
            c.montant_paye,
            max(reste, 0),
            c.get_statut_display(),
            c.date_paiement.strftime("%d/%m/%Y") if c.date_paiement else "",
            c.mode_paiement or "",
            c.modifie_par.username if c.modifie_par else "",
            c.date_modification_paiement.strftime("%d/%m/%Y %H:%M") if c.date_modification_paiement else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cotisations_mensuelles_detail.xlsx"'
    wb.save(response)
    return response


@login_required
def export_cotisations_detail_pdf(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "cotisations_export"):
        return HttpResponseForbidden("Accès refusé.")

    qs = CotisationAssociationMensuelle.objects.filter(
        affiliation__federation=federation
    ).select_related(
        "affiliation__association__entite"
    )

    qs = filtrer_cotisations_mensuelles(qs, request)

    if not qs.exists():
        return export_vide(request, "Aucune cotisation trouvée pour cet export.")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cotisations_mensuelles_detail.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = [
        Paragraph("Détail des cotisations mensuelles", styles["Title"]),
        Spacer(1, 12),
    ]

    data = [[
        "Association",
        "Code",
        "Période",
        "Attendu",
        "Payé",
        "Reste",
        "Statut",
        "Date paiement",
    ]]

    for c in qs.order_by("-annee", "-mois"):
        reste = c.montant_attendu - c.montant_paye

        data.append([
            c.affiliation.association.entite.nom,
            c.affiliation.association.code_paiement or "",
            f"{c.mois}/{c.annee}",
            str(c.montant_attendu),
            str(c.montant_paye),
            str(max(reste, 0)),
            c.get_statut_display(),
            c.date_paiement.strftime("%d/%m/%Y") if c.date_paiement else "",
        ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


@login_required
def export_contributions_deces_excel(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "rapatriements_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = ContributionDecesAssociation.objects.filter(
        dossier__affiliation__federation=federation
    ).select_related(
        "association__entite",
        "dossier__personne",
        "dossier__affiliation__association__entite",
    )

    qs = filtrer_contributions_deces(qs, request)

    if not qs.exists():
        return export_vide(request, "Aucune cotisation décès trouvée pour cet export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotisations décès"

    ws.append([
        "Association",
        "Identifiant",
        "Défunt",
        "Association du défunt",
        "Attendu",
        "Payé",
        "Reste",
        "Statut",
        "Date paiement",
    ])

    for c in qs.order_by("association__entite__nom"):
        reste = c.montant_attendu - c.montant_paye

        ws.append([
            c.association.entite.nom,
            c.association.code_paiement,
            str(c.dossier.personne),
            c.dossier.affiliation.association.entite.nom,
            c.montant_attendu,
            c.montant_paye,
            max(reste, 0),
            c.get_statut_display(),
            c.date_paiement.strftime("%d/%m/%Y") if c.date_paiement else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cotisations_deces.xlsx"'
    wb.save(response)
    return response


@login_required
def export_contributions_deces_pdf(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "rapatriements_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = ContributionDecesAssociation.objects.filter(
        dossier__affiliation__federation=federation
    ).select_related(
        "association__entite",
        "dossier__personne",
        "dossier__affiliation__association__entite",
    )

    qs = filtrer_contributions_deces(qs, request)

    if not qs.exists():
        return export_vide(request, "Aucune cotisation décès trouvée pour cet export.")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cotisations_deces.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = [
        Paragraph("Cotisations décès", styles["Title"]),
        Spacer(1, 12),
    ]

    data = [[
        "Association",
        "Défunt",
        "Attendu",
        "Payé",
        "Reste",
        "Statut",
        "Date paiement",
    ]]

    for c in qs.order_by("association__entite__nom"):
        reste = c.montant_attendu - c.montant_paye

        data.append([
            c.association.entite.nom,
            str(c.dossier.personne),
            str(c.montant_attendu),
            str(c.montant_paye),
            str(max(reste, 0)),
            c.get_statut_display(),
            c.date_paiement.strftime("%d/%m/%Y") if c.date_paiement else "",
        ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


@login_required
def export_personnes_excel(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "personnes_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = Personne.objects.filter(
        adhesions__association__federation=federation
    ).prefetch_related(
        "adhesions__association__entite"
    ).distinct()

    qs = filtrer_membres(qs, request)

    if not qs.exists():
        return export_vide(request, "Aucun membre trouvé pour cet export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"

    ws.append([
        "Numéro",
        "Nom",
        "Prénom",
        "Association",
        "Statut",
        "Téléphone",
        "Email",
    ])

    for p in qs.order_by("nom", "prenom"):
        associations = ", ".join(
            adhesion.association.entite.nom
            for adhesion in p.adhesions.all()
        )

        ws.append([
            p.numero,
            p.nom,
            p.prenom,
            associations,
            p.statut,
            p.telephone,
            p.email,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="membres.xlsx"'
    wb.save(response)
    return response


@login_required
def export_personnes_pdf(request):
    federations = get_user_entites(request.user).filter(type="federation")
    federation = federations.first()

    if not federation or not user_has_permission(request.user, federation, "personnes_view"):
        return HttpResponseForbidden("Accès refusé.")

    qs = Personne.objects.filter(
        adhesions__association__federation=federation
    ).prefetch_related(
        "adhesions__association__entite"
    ).distinct()

    qs = filtrer_membres(qs, request)

    if not qs.exists():
        return export_vide(request, "Aucun membre trouvé pour cet export.")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="membres.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = [
        Paragraph("Liste des membres", styles["Title"]),
        Spacer(1, 12),
    ]

    data = [[
        "Numéro",
        "Nom",
        "Prénom",
        "Association",
        "Statut",
        "Téléphone",
    ]]

    for p in qs.order_by("nom", "prenom"):
        associations = ", ".join(
            adhesion.association.entite.nom
            for adhesion in p.adhesions.all()
        )

        data.append([
            p.numero or "",
            p.nom,
            p.prenom,
            associations,
            p.statut,
            p.telephone or "",
        ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    return response